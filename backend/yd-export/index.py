"""Экспорт кампаний Яндекс Директ в XLSX формат Директ Коммандера.
Генерирует файл с листами «Объявления», «Фразы» по образцу ЯД.
"""
import json
import os
import io
import base64
import psycopg2
from psycopg2.extras import RealDictCursor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


CORS_HEADERS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Methods': 'GET, POST, OPTIONS',
    'Access-Control-Allow-Headers': 'Content-Type, X-Auth-Token',
    'Access-Control-Max-Age': '86400',
}


def _resp(status, body):
    return {
        'statusCode': status,
        'headers': {**CORS_HEADERS, 'Content-Type': 'application/json'},
        'isBase64Encoded': False,
        'body': json.dumps(body, default=str),
    }


def _user_by_token(cur, token):
    if not token:
        return None
    safe = token.replace("'", "''")
    cur.execute(
        f"SELECT u.id, u.email FROM users u "
        f"JOIN sessions s ON s.user_id = u.id "
        f"WHERE s.token = '{safe}' AND s.expires_at > NOW() LIMIT 1"
    )
    return cur.fetchone()


# Маппинг типов кампаний на читаемые имена в ЯД Коммандере
CAMPAIGN_TYPE_NAME = {
    'text': 'Текстово-графические объявления',
    'network': 'Реклама в сетях (РСЯ)',
    'master': 'Мастер кампаний',
}

STRATEGY_NAME = {
    'manual_cpc': 'Ручное управление ставками',
    'max_clicks': 'Максимум кликов',
    'max_clicks_with_budget': 'Максимум кликов с недельным бюджетом',
    'max_conversions': 'Максимум конверсий',
    'target_cpa': 'Оплата за конверсии',
    'target_roas': 'Целевая ДРР',
}

MATCH_TYPE_PREFIX = {
    'broad': '',
    'phrase': '"',
    'exact': '!',
}


def _wrap_keyword(phrase: str, match_type: str) -> str:
    """Преобразует фразу к синтаксису Директа: фразовое в кавычки, точное с !."""
    p = (phrase or '').strip()
    if not p:
        return p
    if match_type == 'phrase':
        return f'"{p}"'
    if match_type == 'exact':
        return f'!{p}'
    return p


def _style_header(ws, row_idx, ncols):
    fill = PatternFill('solid', fgColor='FFFFCB05')  # жёлтый ЯД
    font = Font(bold=True, color='FF333333')
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical='center', wrap_text=True)


def _generate_xlsx(campaign):
    """Генерирует XLSX по структуре, понятной Директ Коммандеру.
    Минимально совместимый формат: лист «Объявления» с базовыми колонками + лист «Ключевые фразы»."""
    wb = Workbook()

    # ============ Лист 1: Объявления ============
    ws = wb.active
    ws.title = 'Объявления'

    headers = [
        'Доп. объявление', 'Доп. объявление группы',
        'ID кампании', 'Название кампании', 'Тип кампании',
        'Стратегия показа', 'Дневной бюджет',
        'ID группы', 'Название группы', 'Регионы показа',
        'Минус-фразы', 'Заголовок 1', 'Заголовок 2',
        'Текст', 'Длина текста', 'Ссылка', 'Отображаемая ссылка',
        'Быстрые ссылки (заголовки)', 'Быстрые ссылки (адреса)',
        'Уточнения',
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.row_dimensions[1].height = 32

    # Регионы — берём из jsonb
    regions = campaign.get('regions') or []
    if isinstance(regions, str):
        try:
            regions = json.loads(regions)
        except Exception:
            regions = []
    regions_str = ', '.join(regions) if regions else 'Россия'

    # Минус-фразы кампании
    neg = (campaign.get('negative_keywords') or '').strip()

    cid = campaign['id']
    cname = campaign.get('name') or f'Кампания #{cid}'
    ctype = CAMPAIGN_TYPE_NAME.get(campaign.get('campaign_type'), 'Текстово-графические объявления')
    strategy = STRATEGY_NAME.get(campaign.get('strategy_type'), 'Ручное управление ставками')
    daily = float(campaign.get('daily_budget') or 0)

    for group in (campaign.get('groups') or []):
        gid = group.get('id') or ''
        gname = group.get('name') or 'Группа'
        ads = group.get('ads') or []
        if not ads:
            ads = [{
                'title1': '', 'title2': '', 'body': '', 'href': '',
                'display_url': '', 'sitelinks': [], 'callouts': [],
            }]

        for ai, ad in enumerate(ads):
            sitelinks = ad.get('sitelinks') or []
            sl_titles = ' | '.join((s.get('title') or '').strip() for s in sitelinks if s.get('title'))
            sl_urls = ' | '.join((s.get('url') or '').strip() for s in sitelinks if s.get('url'))
            callouts = ad.get('callouts') or []
            callout_str = ' | '.join((c.get('text') or '').strip() for c in callouts if c.get('text'))

            body_text = ad.get('body') or ''
            row = [
                '+' if ai > 0 else '-',  # Доп. объявление (вторые/третьи в группе помечаются +)
                '-',
                cid,
                cname,
                ctype,
                strategy,
                f"{daily:.2f}" if ai == 0 else '',
                gid,
                gname,
                regions_str if ai == 0 else '',
                neg if ai == 0 else '',
                ad.get('title1') or '',
                ad.get('title2') or '',
                body_text,
                len(body_text),
                ad.get('href') or '',
                ad.get('display_url') or '',
                sl_titles,
                sl_urls,
                callout_str,
            ]
            ws.append(row)

    # Ширина колонок
    widths = [12, 12, 12, 28, 22, 26, 14, 12, 22, 22, 28, 30, 22, 40, 12, 28, 22, 30, 30, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = 'A2'

    # ============ Лист 2: Ключевые фразы ============
    ws2 = wb.create_sheet('Фразы')
    headers2 = [
        'ID кампании', 'Название кампании', 'ID группы', 'Название группы',
        'Фраза', 'Тип соответствия', 'Ставка', 'Минус-фразы фразы',
    ]
    ws2.append(headers2)
    _style_header(ws2, 1, len(headers2))
    ws2.row_dimensions[1].height = 32

    for group in (campaign.get('groups') or []):
        gid = group.get('id') or ''
        gname = group.get('name') or 'Группа'
        keywords = group.get('keywords') or []
        for kw in keywords:
            phrase = kw.get('phrase') or ''
            match_type = kw.get('match_type') or 'broad'
            bid = float(kw.get('bid') or 0)
            row = [
                cid, cname, gid, gname,
                _wrap_keyword(phrase, match_type),
                {'broad': 'Широкое', 'phrase': 'Фразовое', 'exact': 'Точное'}.get(match_type, 'Широкое'),
                f"{bid:.2f}" if bid > 0 else 'авто',
                '',
            ]
            ws2.append(row)

    widths2 = [12, 28, 12, 22, 50, 18, 12, 28]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w
    ws2.freeze_panes = 'A2'

    # ============ Лист 3: Сводка ============
    ws3 = wb.create_sheet('Сводка')
    summary_rows = [
        ['Параметр', 'Значение'],
        ['Название кампании', cname],
        ['Тип кампании', ctype],
        ['Стратегия', strategy],
        ['Дневной бюджет, ₽', daily],
        ['Недельный бюджет, ₽', float(campaign.get('weekly_budget') or 0)],
        ['Регионы показа', regions_str],
        ['Минус-фразы кампании', neg],
        ['Счётчик Метрики', campaign.get('counter_id') or '—'],
        ['Цели Метрики', campaign.get('counter_goals') or '—'],
        ['UTM-шаблон', campaign.get('utm_template') or '—'],
        ['Групп', len(campaign.get('groups') or [])],
        ['Объявлений',
         sum(len(g.get('ads') or []) for g in (campaign.get('groups') or []))],
        ['Ключевых фраз',
         sum(len(g.get('keywords') or []) for g in (campaign.get('groups') or []))],
        ['Заметки', campaign.get('notes') or ''],
    ]
    for r in summary_rows:
        ws3.append(r)
    _style_header(ws3, 1, 2)
    ws3.column_dimensions['A'].width = 28
    ws3.column_dimensions['B'].width = 60
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Сохраняем в буфер
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _load_full_campaign(cur, user_id, campaign_id):
    cur.execute(f"""
        SELECT * FROM yd_campaigns WHERE id = {int(campaign_id)} AND user_id = {int(user_id)} LIMIT 1
    """)
    c = cur.fetchone()
    if not c:
        return None
    cid = c['id']
    cur.execute(f"SELECT * FROM yd_ad_groups WHERE campaign_id = {int(cid)} ORDER BY sort_order, id")
    groups = [dict(r) for r in cur.fetchall()]
    group_ids = [g['id'] for g in groups]
    ads_by_group = {gid: [] for gid in group_ids}
    keywords_by_group = {gid: [] for gid in group_ids}
    if group_ids:
        ids_sql = ','.join(str(int(g)) for g in group_ids)
        cur.execute(f"SELECT * FROM yd_ads WHERE group_id IN ({ids_sql}) ORDER BY sort_order, id")
        for r in cur.fetchall():
            ads_by_group.setdefault(r['group_id'], []).append(dict(r))
        cur.execute(f"SELECT * FROM yd_keywords WHERE group_id IN ({ids_sql}) ORDER BY id")
        for r in cur.fetchall():
            keywords_by_group.setdefault(r['group_id'], []).append(dict(r))
    for g in groups:
        g['ads'] = ads_by_group.get(g['id'], [])
        g['keywords'] = keywords_by_group.get(g['id'], [])
    out = dict(c)
    out['groups'] = groups
    return out


def _export_campaign(cur, user_id, body):
    """Экспорт одной кампании. Возвращает base64 файла."""
    cid = int(body.get('id') or 0)
    if not cid:
        return {'error': 'id required'}
    campaign = _load_full_campaign(cur, user_id, cid)
    if not campaign:
        return {'error': 'not found'}

    xlsx_bytes = _generate_xlsx(campaign)
    b64 = base64.b64encode(xlsx_bytes).decode('ascii')

    # Обновляем статус кампании
    cur.execute(f"""
        UPDATE yd_campaigns SET status = 'exported', updated_at = NOW()
        WHERE id = {cid} AND user_id = {int(user_id)}
    """)

    name = (campaign.get('name') or f'campaign_{cid}').replace(' ', '_')[:60]
    return {
        'ok': True,
        'filename': f'{name}.xlsx',
        'content_base64': b64,
        'size_bytes': len(xlsx_bytes),
    }


def _preview_campaign(cur, user_id, campaign_id):
    """Превью того, что попадёт в файл — без генерации XLSX."""
    campaign = _load_full_campaign(cur, user_id, campaign_id)
    if not campaign:
        return {'error': 'not found'}

    issues = []
    groups = campaign.get('groups') or []
    if not groups:
        issues.append({'type': 'error', 'msg': 'Нет ни одной группы объявлений'})

    total_ads = 0
    total_keywords = 0
    for g in groups:
        ads = g.get('ads') or []
        keywords = g.get('keywords') or []
        total_ads += len(ads)
        total_keywords += len(keywords)

        if not ads:
            issues.append({'type': 'warning', 'msg': f'Группа «{g.get("name")}» — нет объявлений'})
        for ai, ad in enumerate(ads):
            if not ad.get('title1'):
                issues.append({'type': 'error', 'msg': f'«{g.get("name")}» / Объявление {ai+1}: пустой заголовок 1'})
            if not ad.get('body'):
                issues.append({'type': 'warning', 'msg': f'«{g.get("name")}» / Объявление {ai+1}: пустой текст'})
            if len(ad.get('title1') or '') > 56:
                issues.append({'type': 'error', 'msg': f'«{g.get("name")}» / Объявление {ai+1}: заголовок 1 длиннее 56'})
            if len(ad.get('body') or '') > 81:
                issues.append({'type': 'error', 'msg': f'«{g.get("name")}» / Объявление {ai+1}: текст длиннее 81'})
            if not ad.get('href'):
                issues.append({'type': 'warning', 'msg': f'«{g.get("name")}» / Объявление {ai+1}: нет посадочной ссылки'})

        if campaign.get('campaign_type') != 'master' and not keywords:
            issues.append({'type': 'warning', 'msg': f'«{g.get("name")}» — нет ключевых фраз'})

    return {
        'campaign_name': campaign.get('name'),
        'campaign_type': campaign.get('campaign_type'),
        'groups_count': len(groups),
        'ads_count': total_ads,
        'keywords_count': total_keywords,
        'issues': issues,
        'ready_to_export': not any(i['type'] == 'error' for i in issues),
    }


def handler(event, context):
    """Экспорт кампаний ЯД в XLSX (формат Директ Коммандера)"""
    method = event.get('httpMethod', 'GET')
    if method == 'OPTIONS':
        return {'statusCode': 200, 'headers': CORS_HEADERS, 'isBase64Encoded': False, 'body': ''}

    headers = event.get('headers') or {}
    headers_lower = {k.lower(): v for k, v in headers.items()}
    token = headers_lower.get('x-auth-token', '')

    params = event.get('queryStringParameters') or {}
    action = params.get('action', '')

    body_raw = event.get('body') or '{}'
    try:
        body = json.loads(body_raw) if body_raw else {}
    except Exception:
        body = {}

    dsn = os.environ.get('DATABASE_URL')
    if not dsn:
        return _resp(500, {'error': 'DATABASE_URL is not configured'})

    conn = psycopg2.connect(dsn)
    conn.autocommit = True
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        user = _user_by_token(cur, token)
        if not user:
            return _resp(401, {'error': 'Не авторизован'})
        uid = user['id']

        if method == 'GET' and action == 'preview':
            cid = params.get('id') or 0
            return _resp(200, _preview_campaign(cur, uid, cid))
        if method == 'POST' and action == 'export':
            return _resp(200, _export_campaign(cur, uid, body))

        return _resp(400, {'error': 'Неизвестное действие'})
    except Exception as e:
        import traceback, sys
        print(f"[YD-EXPORT ERROR] {e}\n{traceback.format_exc()}", file=sys.stderr, flush=True)
        return _resp(500, {'error': str(e)[:500]})
    finally:
        cur.close()
        conn.close()
